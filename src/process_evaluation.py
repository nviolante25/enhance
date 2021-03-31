import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook


def reference_results(worksheet, evaluated_methods):
    def get_comment(worksheet, num_tests):
        comment = [str(worksheet[int(7+scene_idx)][2+i].comment).split(':')[-1] for i in range(num_tests)]
        comment = ['' if cmt == 'None' else cmt.split('\n')[0] for cmt in comment]
        return comment
    
    results = pd.DataFrame(worksheet.values).iloc[6:, 1:11]
    results.columns = ['Scenes'] + evaluated_methods
    results = results.set_index('Scenes')
    comments = results.copy()
    comments.iloc[:,:] = ''
    
    num_scenes, num_tests = results.shape
    for scene in results.index:
        # Get ordered methods for the scene
        idx = methods_reference.loc[methods_reference[0].str.contains(scene)==True].index[0]
        scene_methods = methods_reference.iloc[idx+1:idx+1+num_tests]
        methods = [methods_reference.iloc[idx+1:idx+1+num_tests].iloc[i][0].split(':')[0] for i in range(scene_methods.shape[0])]

        scene_idx = np.where(results.index.str.contains(scene) == True)[0][0]
        comment = get_comment(worksheet, num_tests)
        # Add scene scores for each method
        methods_scores = np.array(results.loc[scene], dtype=np.int8)
        for j, method in enumerate(methods):
            results[method][scene] = methods_scores[j]
            comments[method][scene] = comment[j]
    return results, comments


def count_scores(results):
    df = pd.DataFrame(index=range(3), columns=range(results.shape[1]))
    df.columns = results.columns
    for e in results:
        df[e][0] = results[e].value_counts().get(-1,0)
        df[e][1] = results[e].value_counts().get(0,0)
        df[e][2] = results[e].value_counts().get(1,0)
    df = df.transpose()
    df.columns = [-1, 0, 1]
    return df
    
    
def save_bests(results):
    for i, (scene_name, methods_score) in enumerate(results.iterrows()):
        max_idx = methods_score.astype(np.int8).argmax()
        best_method = methods_score.index[max_idx]
        print(f'Scene {scene_name}, best method {best_method}')
    

if __name__ == '__main__':
    excel = load_workbook('/home/nviolante/projects/retinex/eval_results/Tests.xlsx')
    methods_reference = pd.read_csv('/home/nviolante/projects/retinex/eval_results/reference.txt', header=None)
    methods_reference = methods_reference.replace(r'\s', '', regex=True)
    evaluated_methods = sorted(['quicklook', 'retinex_hsv_02', 'ace_01', 'quicklook_ps', 'log_stretch', 
                                'ace_00', 'retinex_hsv_05', 'quicklook_v2', 'quicklook_ps_ev'])

    evaluator = 'Juanfer'
    evaluators = excel.sheetnames[1:] # exclude reference sheet
    results, comments = reference_results(excel[evaluator], evaluated_methods)
    out = {evaluator: reference_results(excel[evaluator], evaluated_methods) for evaluator in evaluators}
    
    juanfer = count_scores(out['Juanfer'][0])
    gabriel = count_scores(out['Gabriel'][0])
    juanfer.plot.bar(ylim=[0,6])
    gabriel.plot.bar(ylim=[0,6])
    
    save_bests(out['Juanfer'][0])
    save_bests(out['Gabriel'][0])
    
    overall_scenes = pd.concat([out['Juanfer'][0].mean(axis=1), out['Gabriel'][0].mean(axis=1)], axis=1)
    overall_scenes.columns = ['Juanfer', 'Gabriel']
    overall_scenes.plot.barh()
    plt.show()
    print()
